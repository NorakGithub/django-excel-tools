import unittest
from unittest import skip

from openpyxl import Workbook

from django_excel_tools import serializers
from django_excel_tools.exceptions import FieldNotExist, ValidationError
from django_excel_tools.serializers import SerializerMeta


class TestSerializerMeta(unittest.TestCase):
    def test_failed_when_class_meta_is_none(self):
        with self.assertRaises(AssertionError):
            SerializerMeta(None)

    def test_start_index_validation(self):
        with self.assertRaises(AssertionError):
            class Meta:
                pass
            SerializerMeta(Meta)

        with self.assertRaises(AssertionError):
            class Meta:
                start_index = 0
            SerializerMeta(Meta)

        with self.assertRaises(AssertionError):
            class Meta:
                start_index = 'A'
            SerializerMeta(Meta)

    def test_fields_validation(self):
        with self.assertRaises(AssertionError):
            class Meta:
                start_index = 1
            SerializerMeta(Meta)

        with self.assertRaises(AssertionError):
            class Meta:
                start_index = 1
                fields = None
            SerializerMeta(Meta)

        with self.assertRaises(AssertionError):
            class Meta:
                start_index = 1
                fields = 1
            SerializerMeta(Meta)

    def test_enable_transaction_validation(self):
        with self.assertRaises(AssertionError):
            class Meta:
                start_index = 1
                fields = ('field1', 'field2')
                enable_transaction = 'A'
            SerializerMeta(Meta)

    def test_validation_should_passed(self):
        class Meta:
            start_index = 1
            fields = ('field1', 'field2')

        meta = SerializerMeta(Meta)
        self.assertEqual(meta.start_index, 1)
        self.assertEqual(meta.fields, ('field1', 'field2'))
        self.assertEqual(meta.enable_transaction, True)

        class Meta:
            start_index = 1
            fields = ('field1', 'field2')
            enable_transaction = False

        meta = SerializerMeta(Meta)
        self.assertFalse(meta.enable_transaction)


class TestSerializer(unittest.TestCase):
    def setUp(self):
        workbook = Workbook()
        self.worksheet = workbook.active
        self.worksheet.append(['Field name 1', 'Field name 2'])
        self.worksheet.append(['value 1', 'value 2'])

    def test_should_field_when_no_class_meta(self):
        class Serializer(serializers.ExcelSerializer):
            pass

        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_not_adding_start_index(self):
        class Serializer(serializers.ExcelSerializer):
            class Meta:
                pass
        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_not_adding_fields(self):
        class Serializer(serializers.ExcelSerializer):
            class Meta:
                start_index = 1

        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_no_fields(self):
        class Serializer(serializers.ExcelSerializer):
            class Meta:
                start_index = 1
        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_fields_is_not_set(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1

        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_field_is_not_defined_included(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name')

        with self.assertRaises(FieldNotExist):
            Serializer(self.worksheet)

    def test_should_call_row_extra_validation_when_override(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

            def row_extra_validation(self, index, cleaned_row):
                assert cleaned_row['field_name_1'] == 'value 1'
                assert cleaned_row['field_name_2'] == 'value 2'

        Serializer(self.worksheet)

    @skip
    def test_should_fail_if_sheet_has_more_fields_than_serializer(self):
        self.worksheet['C1'] = 'Field Name 3'
        self.worksheet['C2'] = 'Value 3'

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

        serializer = Serializer(self.worksheet)
        self.assertEqual(self.worksheet.max_column, 3)
        self.assertNotEqual(self.worksheet.max_column, len(serializer.fields))
        self.assertEqual(serializer.errors[0],
                         'Required 2 fields, but given excel has 3 fields, amount of field should be the same. '
                         '[Tip] You might select the wrong excel format.'
                         )

    def test_should_ignore_if_worksheet_has_blank_excess_columns(self):
        self.worksheet['C1'] = ''
        self.worksheet['D1'] = ''
        self.worksheet['E1'] = ''

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

        serializer = Serializer(self.worksheet)
        self.assertFalse(serializer.validation_errors)

    def test_should_ignore_if_worksheet_has_blank_excess_rows(self):
        self.worksheet['A3'] = ''
        self.worksheet['B3'] = ''
        self.worksheet['A4'] = ''
        self.worksheet['B4'] = ''

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

        serializer = Serializer(self.worksheet)
        self.assertFalse(serializer.validation_errors)

    def test_extra_clean_should_raise_validation_error(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

            def extra_clean_field_name_2(self, value):
                if value == 'value 2':
                    raise ValidationError('This value is not excepted.')

        self.worksheet.append(['value 3', 'value 3'])
        serializer = Serializer(self.worksheet)
        self.assertIsNotNone(serializer.validation_errors)
        self.assertEqual(len(serializer.validation_errors), 1)

    def test_extra_clean_should_change_the_value(self):
        expected_value = 'Succeed'

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

            def extra_clean_field_name_2(self, value):
                if value == 'value 3':
                    raise ValidationError('This value is not excepted.')
                return expected_value

        self.worksheet.append(['value', 'value temp'])
        serializer = Serializer(self.worksheet)
        self.assertEqual(serializer.validation_errors, [])
        self.assertEqual(serializer.cleaned_data[0]['field_name_2'], expected_value)
        self.assertEqual(serializer.cleaned_data[1]['field_name_2'], expected_value)
