#!/usr/bin/env python
# -*- coding: utf-8 -*-
import unittest

import datetime
from openpyxl import Workbook

from django_excel_tools import serializers
from django_excel_tools.exceptions import FieldNotExist, ValidationError
from tests.workbook import WorkbookTesting, OrderExcelSerializer


class TestSerializer(unittest.TestCase):

    def setUp(self):
        workbook = Workbook()
        self.worksheet = workbook.active
        self.worksheet.append(['Field name 1', 'Field name 2'])
        self.worksheet.append(['value 1', 'value 2'])

    def test_should_field_when_no_class_meta(self):
        class Serializer(serializers.ExcelSerializer):
            pass

        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_not_adding_start_index(self):
        class Serializer(serializers.ExcelSerializer):
            class Meta:
                pass
        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_not_adding_fields(self):
        class Serializer(serializers.ExcelSerializer):
            class Meta:
                start_index = 1

        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_no_fields(self):
        class Serializer(serializers.ExcelSerializer):
            class Meta:
                start_index = 1
        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_fields_is_not_set(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1

        with self.assertRaises(AssertionError):
            Serializer(self.worksheet)

    def test_should_failed_when_field_is_not_defined_included(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name')

        with self.assertRaises(FieldNotExist):
            Serializer(self.worksheet)

    def test_should_call_row_extra_validation_when_override(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

            def row_extra_validation(self, index, cleaned_row):
                assert cleaned_row['field_name_1'] == 'value 1'
                assert cleaned_row['field_name_2'] == 'value 2'

        Serializer(self.worksheet)

    def test_should_fail_if_sheet_has_more_fields_than_serializer(self):
        self.worksheet['C1'] = 'Field Name 3'
        self.worksheet['C2'] = 'Value 3'

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

        serializer = Serializer(self.worksheet)
        self.assertEqual(self.worksheet.max_column, 3)
        self.assertNotEqual(self.worksheet.max_column, len(serializer.fields))
        self.assertEqual(serializer.errors[0],
                         'Required 2 fields, but given excel has 3 fields, amount of field should be the same. '
                         '[Tip] You might select the wrong excel format.'
                         )

    def test_should_ignore_if_worksheet_has_blank_excess_columns(self):
        self.worksheet['C1'] = ''
        self.worksheet['D1'] = ''
        self.worksheet['E1'] = ''

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

        serializer = Serializer(self.worksheet)
        self.assertFalse(serializer.errors)

    def test_should_ignore_if_worksheet_has_blank_excess_rows(self):
        self.worksheet['A3'] = ''
        self.worksheet['B3'] = ''
        self.worksheet['A4'] = ''
        self.worksheet['B4'] = ''

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10, verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10, verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

        serializer = Serializer(self.worksheet)
        self.assertFalse(serializer.errors)

    def test_extra_clean_should_raise_validation_error(self):
        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

            def extra_clean_field_name_2(self, value):
                if value == 'value 2':
                    raise ValidationError('This value is not excepted.')

        self.worksheet.append(['value 3', 'value 3'])
        serializer = Serializer(self.worksheet)
        self.assertIsNotNone(serializer.errors)
        self.assertEqual(len(serializer.errors), 1)

    def test_extra_clean_should_change_the_value(self):
        expected_value = 'Succeed'

        class Serializer(serializers.ExcelSerializer):
            field_name_1 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 1')
            field_name_2 = serializers.CharField(max_length=10,
                                                 verbose_name='Field name 2')

            class Meta:
                start_index = 1
                fields = ('field_name_1', 'field_name_2')

            def extra_clean_field_name_2(self, value):
                if value == 'value 3':
                    raise ValidationError('This value is not excepted.')
                return expected_value

        self.worksheet.append(['value', 'value temp'])
        serializer = Serializer(self.worksheet)
        self.assertEqual(serializer.errors, [])
        self.assertEqual(serializer.cleaned_data[0]['field_name_2'], expected_value)
        self.assertEqual(serializer.cleaned_data[1]['field_name_2'], expected_value)


class TestField(unittest.TestCase):

    def setUp(self):
        workbook = WorkbookTesting()
        self.worksheet = workbook.worksheet
        self.basic_data = ['Shop', '10', '2017-07-07', 100, '20180101', 201801, '100', u'無', 'AB', '123/Home', 'Yes']

    def test_allow_blank_with_empty_string(self):
        self.basic_data[9] = ''
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)

    def test_allow_blank_but_there_is_data(self):
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)
        self.assertEqual(serializer.cleaned_data[0]['shop_name'], 'Shop')

    def test_not_allow_blank_with_empty_string(self):
        self.basic_data[0] = '  '  # Set shop field to blank
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertTrue(serializer.errors)

    def test_not_allow_blank_with_zero(self):
        self.basic_data[0] = 0
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)
        self.assertEqual(serializer.cleaned_data[0]['shop_name'], '0')

    def test_max_length(self):
        self.basic_data[0] = '123456789010'
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertTrue(serializer.errors)

    def test_default_int(self):
        self.basic_data[6] = ''
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)
        self.assertEqual(serializer.cleaned_data[0]['weight'], 0)

    def test_datetime_field(self):
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)

        date = datetime.datetime(year=2018, month=1, day=1).date()
        self.assertEqual(serializer.cleaned_data[0]['registered_date'], date)

    def test_datetime_blank_with_none(self):
        self.basic_data[5] = None
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)

    def test_datetime_blank_with_empty_str(self):
        self.basic_data[5] = ''
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)

    def test_datetime_blank_should_receive_error(self):
        self.basic_data[2] = ''
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertTrue(serializer.errors)

    def test_invalid_integer(self):
        self.basic_data[3] = 'Not Integer'
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertTrue(serializer.errors)

    def test_case_sensitive(self):
        self.basic_data[10] = 'yes'
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertFalse(serializer.errors)

    def test_invalid_integer_with_default_value(self):
        self.basic_data[6] = 'Not Integer'
        self.worksheet.append(self.basic_data)
        serializer = OrderExcelSerializer(self.worksheet)
        self.assertTrue(serializer.errors)


class TestResult(unittest.TestCase):

    def setUp(self):
        workbook = WorkbookTesting()
        self.worksheet = workbook.worksheet
        self.worksheet.append(workbook.row_1)
        self.worksheet.append(workbook.row_2)

    def test_result(self):
        order_serializer = OrderExcelSerializer(self.worksheet)
        self.assertEqual(order_serializer.errors, [], order_serializer.errors)
        data = order_serializer.cleaned_data
        first_row = data[0]
        self.assertEqual(first_row['shop_name'], 'Shop A')
        self.assertEqual(first_row['quantity'], 100)
        self.assertEqual(first_row['default_checked'], True)
        self.assertEqual(first_row['weight'], 100)
        self.assertEqual(first_row['address'], '123/Home')

        second_row = data[1]
        self.assertEqual(second_row['shop_name'], 'Shop B')
        self.assertEqual(second_row['quantity'], 1000)
        self.assertEqual(second_row['default_checked'], False)
        self.assertEqual(second_row['weight'], 0)
        self.assertEqual(second_row['address'], '')
